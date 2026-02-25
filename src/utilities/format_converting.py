"""
Format Converting Module

Converts files in-place to standardized formats, deletes the original,
and returns True/False for success/failure.

Supported conversions:
	Images       → PNG   (Pillow, rawpy, pillow-heif)
	PNG          → PDF   (Pillow + ReportLab)
	Video        → MP4   (FFmpeg: H.264 + AAC)
	Audio        → MP3   (FFmpeg: libmp3lame)
	Documents    → PDF   (LibreOffice headless)
	HTML         → PDF   (WeasyPrint)
	Email        → PDF   (HTML render → WeasyPrint)
	XLSX         → CSV   (openpyxl)
	Publisher    → PDF   (LibreOffice headless)

Author: Ashiq Gazi
"""

from __future__ import annotations

import csv
import email
import email.policy
import logging
import mailbox
import platform
import shutil
import subprocess
from email.header import decode_header , make_header
from email.message import Message
from pathlib import Path
from typing import Iterator


# ══════════════════════════════════════════════════════════════════════════════
# INTERNAL HELPERS
# ══════════════════════════════════════════════════════════════════════════════


def _run( args: list[ str ] , logger: logging.Logger , timeout: int = 600 ) -> tuple[ bool , str ] :
	"""
	Execute an external command and capture its output.

	Args:
		args:    Command and arguments as a list of strings (e.g. ["ffmpeg", "-y", ...]).
		logger:  Logger instance for debug/warning/error messages.
		timeout: Maximum seconds the process is allowed to run before being killed.

	Returns:
		A tuple of (success: bool, combined_stdout_and_stderr: str).
	"""
	cmd_str = " ".join( str( a ) for a in args )
	logger.debug( "Executing command: %s" , cmd_str )

	try :
		result = subprocess.run( args , capture_output=True , text=True , timeout=timeout , check=False )
		output = (result.stdout or "") + (result.stderr or "")

		if result.returncode != 0 :
			logger.warning(
					"Command exited with non-zero code %d — executable: '%s' — output (first 600 chars): %s" ,
					result.returncode , args[ 0 ] , output[ :600 ] ,
			)
			return False , output

		logger.debug( "Command completed successfully: %s" , args[ 0 ] )
		return True , output

	except subprocess.TimeoutExpired :
		logger.error( "Command timed out after %d seconds: %s" , timeout , cmd_str )
		return False , "timeout"

	except FileNotFoundError :
		logger.error( "Executable not found: '%s' — is it installed and on PATH?" , args[ 0 ] )
		return False , f"executable not found: {args[ 0 ]}"

	except Exception as exc :
		logger.error( "Unexpected error running '%s': %s" , args[ 0 ] , exc )
		return False , str( exc )


def _delete_original( src: Path , dst: Path , logger: logging.Logger ) -> None :
	"""
	Remove the original file after a successful conversion.

	Args:
		src:    The original file that was converted.
		dst:    The newly created converted file (logged for context).
		logger: Logger instance.
	"""
	logger.debug( "Removing original file: %s" , src )
	src.unlink( )
	logger.info( "Conversion complete: '%s' → '%s'" , src.name , dst.name )


def _decode_header_value( raw: str | None ) -> str :
	"""
	Decode an RFC-2047 encoded email header into a plain Unicode string.
	Returns an empty string if the header is missing or cannot be decoded.
	"""
	if not raw :
		return ""
	try :
		return str( make_header( decode_header( raw ) ) )
	except Exception :
		return raw or ""


def _find_libreoffice( ) -> str :
	"""
	Locate the LibreOffice 'soffice' executable on the current system.

	On Windows, checks the two standard install paths first, then falls back to PATH.
	On Linux/macOS, checks PATH for 'soffice' and 'libreoffice'.

	Returns:
		The full path to the soffice executable.

	Raises:
		RuntimeError: If LibreOffice cannot be found anywhere.
	"""
	import os

	if platform.system( ) == "Windows" :
		# Windows: LibreOffice is rarely on PATH — check default install locations
		candidates = [
			r"C:\Program Files\LibreOffice\program\soffice.exe" ,
			r"C:\Program Files (x86)\LibreOffice\program\soffice.exe" ,
		]
		for path in candidates :
			if os.path.isfile( path ) :
				return path

		# Last resort: maybe the user added it to PATH manually
		found = shutil.which( "soffice" )
		if found :
			return found

		raise RuntimeError(
				"LibreOffice not found. Expected at "
				r"'C:\Program Files\LibreOffice\program\soffice.exe'. "
				"Download from https://www.libreoffice.org/download/" ,
		)

	# Linux / macOS: check PATH for common executable names
	for cmd in ("soffice" , "libreoffice") :
		found = shutil.which( cmd )
		if found :
			return found

	raise RuntimeError(
			"LibreOffice not found on PATH. "
			"Install with: sudo apt install libreoffice  OR  brew install libreoffice" ,
	)


# ══════════════════════════════════════════════════════════════════════════════
# IMAGE → PNG
# ══════════════════════════════════════════════════════════════════════════════


def convert_image_to_png( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert any supported image to PNG format, in-place.

	Handles:
		- NEF (Nikon RAW) via the 'rawpy' library
		- HEIC (Apple) via the 'pillow-heif' library
		- Everything else (JPEG, BMP, TIFF, WebP, GIF, etc.) via Pillow

	The original file is deleted on success.

	Args:
		src:    Path to the source image file.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".png" )
	suffix = src.suffix.lower( )
	logger.info( "Starting image→PNG conversion: '%s' (%s)" , src.name , suffix )

	try :
		if suffix == ".nef" :
			# Nikon RAW files need the rawpy library to demosaic sensor data
			import rawpy
			from PIL import Image

			logger.debug( "Using rawpy to decode NEF sensor data with camera white balance" )
			with rawpy.imread( str( src ) ) as raw :
				rgb = raw.postprocess( use_camera_wb=True , output_bps=8 )
			Image.fromarray( rgb ).save( dst , format="PNG" )

		elif suffix in (".heic" , ".heif") :
			# Apple HEIC/HEIF files require the pillow-heif plugin AND the
			# system-level libheif C library.  We try pillow-heif first; if
			# that fails for any reason we fall back to FFmpeg (which ships
			# its own HEIF decoder when compiled with --enable-libheif).
			from PIL import Image

			heif_ok = False
			try :
				from pillow_heif import register_heif_opener
				# Must register BEFORE the first Image.open() call in this
				# process, otherwise Pillow's plugin registry ignores it.
				register_heif_opener( )
				logger.debug( "pillow-heif registered — attempting Pillow decode" )

				with Image.open( src ) as img :
					# Apply EXIF orientation so the image isn't rotated/flipped
					from PIL import ImageOps
					img = ImageOps.exif_transpose( img )
					img.convert( "RGB" ).save( dst , format="PNG" )

				heif_ok = True
				logger.debug( "HEIC decoded successfully via pillow-heif" )

			except ImportError :
				logger.warning(
						"pillow-heif is not installed (pip install pillow-heif) — "
						"falling back to FFmpeg for HEIC decode" ,
				)
			except Exception as exc :
				# pillow-heif is installed but decoding failed — this usually
				# means the system libheif library is missing, too old, or
				# the file uses an unsupported codec (e.g. HEVC without a
				# license on some builds).
				logger.warning(
						"pillow-heif failed to decode '%s': %s — falling back to FFmpeg" ,
						src.name , exc ,
				)

			if not heif_ok :
				# Fallback: FFmpeg can decode HEIC/HEIF and output PNG directly
				logger.info( "Attempting HEIC→PNG conversion via FFmpeg fallback" )
				ok , output = _run(
						[ "ffmpeg" , "-y" , "-i" , str( src ) , str( dst ) ] ,
						logger ,
						timeout=120 ,
				)
				if not ok or not dst.exists( ) :
					logger.error(
							"Both pillow-heif and FFmpeg failed to convert HEIC file '%s'. "
							"Ensure libheif is installed (apt install libheif-dev / "
							"brew install libheif) or that FFmpeg was built with HEIF support. "
							"FFmpeg output: %s" ,
							src.name , output[ :400 ] ,
					)
					return False

				logger.debug( "HEIC decoded successfully via FFmpeg fallback" )

		else :
			# Standard image formats handled directly by Pillow
			from PIL import Image

			with Image.open( src ) as img :
				mode = img.mode
				# Some modes (e.g. CMYK, P, I) need conversion before saving as PNG
				if mode not in ("RGBA" , "RGB" , "L" , "LA") :
					target = "RGBA" if getattr( img , "has_transparency_data" , False ) else "RGB"
					logger.debug( "Converting pixel mode %s → %s for PNG compatibility" , mode , target )
					img = img.convert( target )
				img.save( dst , format="PNG" )

		size_kb = dst.stat( ).st_size / 1024
		logger.info( "PNG written successfully: '%s' (%.1f KB)" , dst.name , size_kb )
		_delete_original( src , dst , logger )
		return True

	except ImportError as exc :
		logger.error( "Missing dependency for %s conversion: %s" , suffix , exc )
		return False

	except Exception as exc :
		logger.error( "Image→PNG conversion failed for '%s': %s" , src.name , exc )
		# Clean up partial output if it was created
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False


# ══════════════════════════════════════════════════════════════════════════════
# PNG → PDF
# ══════════════════════════════════════════════════════════════════════════════


def convert_png_to_pdf( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert a PNG image to a single-page PDF, in-place.

	Uses Pillow to read the image dimensions and ReportLab to create a PDF
	page sized exactly to the image. The image is embedded at full resolution.

	The original PNG is deleted on success.

	Args:
		src:    Path to the source PNG file.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".pdf" )
	logger.info( "Starting PNG→PDF conversion: '%s'" , src.name )

	try :
		from PIL import Image
		from reportlab.lib.units import inch
		from reportlab.pdfgen import canvas

		# Read image dimensions to set the PDF page size exactly
		with Image.open( src ) as img :
			width_px , height_px = img.size
			dpi = img.info.get( "dpi" , (72 , 72) )
			# Convert pixel dimensions to points (1 point = 1/72 inch)
			dpi_x = dpi[ 0 ] if dpi[ 0 ] > 0 else 72
			dpi_y = dpi[ 1 ] if dpi[ 1 ] > 0 else 72
			width_pt = (width_px / dpi_x) * 72
			height_pt = (height_px / dpi_y) * 72

		logger.debug(
				"Image dimensions: %dx%d px, DPI: %s, PDF page: %.1f x %.1f pt" ,
				width_px , height_px , dpi , width_pt , height_pt ,
		)

		# Create a PDF with a single page sized to the image
		c = canvas.Canvas( str( dst ) , pagesize=(width_pt , height_pt) )
		c.drawImage( str( src ) , 0 , 0 , width=width_pt , height=height_pt )
		c.save( )

		size_kb = dst.stat( ).st_size / 1024
		logger.info( "PDF written successfully: '%s' (%.1f KB)" , dst.name , size_kb )
		_delete_original( src , dst , logger )
		return True

	except ImportError as exc :
		logger.error( "Missing dependency for PNG→PDF: %s (pip install reportlab)" , exc )
		return False

	except Exception as exc :
		logger.error( "PNG→PDF conversion failed for '%s': %s" , src.name , exc )
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False


# ══════════════════════════════════════════════════════════════════════════════
# VIDEO → MP4
# ══════════════════════════════════════════════════════════════════════════════


def convert_video_to_mp4( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert a video file to MP4 (H.264 video + AAC audio) in-place using FFmpeg.

	Encoding settings:
		- Video: libx264, CRF 18 (visually lossless), slow preset for quality
		- Audio: AAC at 192 kbps
		- faststart flag for web streaming compatibility

	The original file is deleted on success.

	Args:
		src:    Path to the source video file.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".mp4" )
	logger.info( "Starting video→MP4 conversion: '%s' (this may take a while for large files)" , src.name )

	ok , output = _run(
			[
				"ffmpeg" , "-y" , "-i" , str( src ) ,
				"-c:v" , "libx264" , "-crf" , "18" , "-preset" , "slow" ,
				"-c:a" , "aac" , "-b:a" , "192k" ,
				"-movflags" , "+faststart" ,
				str( dst ) ,
			] ,
			logger ,
			timeout=3600 ,  # 1-hour timeout for large video files
	)

	if not ok or not dst.exists( ) :
		logger.error( "Video→MP4 conversion failed for '%s': %s" , src.name , output[ :600 ] )
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False

	size_mb = dst.stat( ).st_size / (1024 * 1024)
	logger.info( "MP4 written successfully: '%s' (%.1f MB)" , dst.name , size_mb )
	_delete_original( src , dst , logger )
	return True


# ══════════════════════════════════════════════════════════════════════════════
# AUDIO → MP3
# ══════════════════════════════════════════════════════════════════════════════


def convert_audio_to_mp3( src: Path , logger: logging.Logger , bitrate: str = "320k" ) -> bool :
	"""
	Convert an audio file to MP3 in-place using FFmpeg.

	If the source is already an MP3, the stream is copied without re-encoding
	to avoid quality loss from double-compression.

	Args:
		src:     Path to the source audio file.
		logger:  Logger instance.
		bitrate: Target MP3 bitrate (default "320k" for highest quality).

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".mp3" )
	logger.info( "Starting audio→MP3 conversion: '%s' (bitrate=%s)" , src.name , bitrate )

	if src.suffix.lower( ) == ".mp3" :
		# Already MP3 — just copy the audio stream to avoid re-encoding
		logger.debug( "Source is already MP3 — stream-copying to preserve quality" )
		encode_args = [ "-c:a" , "copy" ]
	else :
		# Encode to MP3 using LAME with ID3v2 tags
		encode_args = [ "-c:a" , "libmp3lame" , "-b:a" , bitrate , "-id3v2_version" , "3" , "-write_id3v1" , "1" ]

	ok , output = _run(
			[ "ffmpeg" , "-y" , "-i" , str( src ) ] + encode_args + [ str( dst ) ] ,
			logger ,
			timeout=600 ,
	)

	if not ok or not dst.exists( ) :
		logger.error( "Audio→MP3 conversion failed for '%s': %s" , src.name , output[ :600 ] )
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False

	size_mb = dst.stat( ).st_size / (1024 * 1024)
	logger.info( "MP3 written successfully: '%s' (%.1f MB)" , dst.name , size_mb )
	_delete_original( src , dst , logger )
	return True


# ══════════════════════════════════════════════════════════════════════════════
# HTML → PDF
# ══════════════════════════════════════════════════════════════════════════════


def convert_html_to_pdf( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert an HTML file to PDF in-place using WeasyPrint.

	WeasyPrint renders the HTML the same way a browser would for printing
	(equivalent to opening the file and pressing Ctrl+P).

	Args:
		src:    Path to the source HTML file.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".pdf" )
	logger.info( "Starting HTML→PDF conversion: '%s'" , src.name )

	try :
		from weasyprint import HTML , CSS

		# Render to PDF with sensible page margins
		HTML( filename=str( src ) ).write_pdf(
				str( dst ) ,
				stylesheets=[ CSS( string="@page { margin: 20mm 15mm; }" ) ] ,
		)

		if not dst.exists( ) :
			logger.error( "WeasyPrint produced no output file for '%s'" , src.name )
			return False

		size_kb = dst.stat( ).st_size / 1024
		logger.info( "PDF written successfully: '%s' (%.1f KB)" , dst.name , size_kb )
		_delete_original( src , dst , logger )
		return True

	except ImportError as exc :
		logger.error( "Missing dependency for HTML→PDF: %s (pip install weasyprint)" , exc )
		return False

	except Exception as exc :
		logger.error( "HTML→PDF conversion failed for '%s': %s" , src.name , exc )
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False


# ══════════════════════════════════════════════════════════════════════════════
# DOCUMENT → PDF  (LibreOffice headless)
# ══════════════════════════════════════════════════════════════════════════════


def convert_document_to_pdf( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert a document (DOCX, XLSX, PPTX, ODT, RTF, etc.) to PDF in-place
	using LibreOffice in headless mode.

	LibreOffice writes its output to a directory you specify; since we want
	in-place conversion, we point --outdir at the source file's own directory,
	then delete the original.

	Args:
		src:    Path to the source document.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".pdf" )
	logger.info( "Starting document→PDF conversion via LibreOffice: '%s'" , src.name )

	try :
		soffice = _find_libreoffice( )
		logger.debug( "Found LibreOffice executable at: %s" , soffice )
	except RuntimeError as exc :
		logger.error( "%s" , exc )
		return False

	# Tell LibreOffice to write the PDF directly into the source file's directory
	out_dir = str( src.parent )

	ok , output = _run(
			[
				soffice ,
				"--headless" , "--nologo" , "--norestore" ,
				"--convert-to" , "pdf" ,
				"--outdir" , out_dir ,
				str( src ) ,
			] ,
			logger ,
			timeout=300 ,
	)

	if not ok or not dst.exists( ) :
		logger.error(
				"LibreOffice conversion failed for '%s' — the file may be password-protected "
				"or corrupt. Output: %s" , src.name , output[ :400 ] ,
		)
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False

	size_kb = dst.stat( ).st_size / 1024
	logger.info( "PDF written successfully: '%s' (%.1f KB)" , dst.name , size_kb )
	_delete_original( src , dst , logger )
	return True


# ══════════════════════════════════════════════════════════════════════════════
# PUBLISHER (.pub) → PDF  (LibreOffice headless)
# ══════════════════════════════════════════════════════════════════════════════


def convert_pub_to_pdf( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert a Microsoft Publisher (.pub) file to PDF in-place using LibreOffice.

	Publisher files are a proprietary Microsoft format. LibreOffice has partial
	support for reading them — complex layouts may not render perfectly, but
	text and basic formatting are generally preserved.

	Args:
		src:    Path to the .pub file.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )

	if src.suffix.lower( ) != ".pub" :
		logger.error( "Expected a .pub file but got '%s'" , src.suffix )
		return False

	logger.info(
			"Starting Publisher→PDF conversion: '%s' (note: complex Publisher "
			"layouts may not render perfectly in LibreOffice)" , src.name ,
	)

	# Delegate to the general document→PDF converter since LibreOffice handles .pub
	return convert_document_to_pdf( src , logger )


# ══════════════════════════════════════════════════════════════════════════════
# XLSX → CSV
# ══════════════════════════════════════════════════════════════════════════════


def convert_xlsx_to_csv( src: Path , logger: logging.Logger , sheet_name: str | None = None ) -> bool :
	"""
	Convert an Excel workbook (.xlsx) to CSV in-place using openpyxl.

	By default, the first (active) sheet is exported. Pass sheet_name to pick
	a specific sheet. If the workbook has multiple sheets and no sheet_name is
	given, a warning is logged so the user knows only the active sheet was used.

	The original .xlsx file is deleted on success.

	Args:
		src:        Path to the source .xlsx file.
		logger:     Logger instance.
		sheet_name: Optional name of the sheet to export. Defaults to the active sheet.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".csv" )
	logger.info( "Starting XLSX→CSV conversion: '%s'" , src.name )

	try :
		from openpyxl import load_workbook

		# read_only=True is faster and uses less memory for large workbooks
		# data_only=True gives us computed cell values instead of formulas
		wb = load_workbook( str( src ) , read_only=True , data_only=True )

		# Pick the requested sheet or fall back to the active sheet
		available_sheets = wb.sheetnames
		logger.debug( "Workbook contains %d sheet(s): %s" , len( available_sheets ) , available_sheets )

		if sheet_name :
			if sheet_name not in available_sheets :
				logger.error(
						"Sheet '%s' not found in workbook. Available sheets: %s" ,
						sheet_name , available_sheets ,
				)
				wb.close( )
				return False
			ws = wb[ sheet_name ]
			logger.debug( "Exporting requested sheet: '%s'" , sheet_name )
		else :
			ws = wb.active
			if len( available_sheets ) > 1 :
				logger.warning(
						"Workbook has %d sheets but no sheet_name was specified — "
						"exporting only the active sheet: '%s'" ,
						len( available_sheets ) , ws.title ,
				)
			logger.debug( "Exporting active sheet: '%s'" , ws.title )

		# Write each row to the CSV file
		row_count = 0
		with open( dst , "w" , newline="" , encoding="utf-8" ) as f :
			writer = csv.writer( f )
			for row in ws.iter_rows( values_only=True ) :
				# Convert None cells to empty strings for cleaner CSV output
				writer.writerow( [ "" if cell is None else cell for cell in row ] )
				row_count += 1

		wb.close( )

		size_kb = dst.stat( ).st_size / 1024
		logger.info(
				"CSV written successfully: '%s' (%d rows, %.1f KB)" ,
				dst.name , row_count , size_kb ,
		)
		_delete_original( src , dst , logger )
		return True

	except ImportError as exc :
		logger.error( "Missing dependency for XLSX→CSV: %s (pip install openpyxl)" , exc )
		return False

	except Exception as exc :
		logger.error( "XLSX→CSV conversion failed for '%s': %s" , src.name , exc )
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False


# ══════════════════════════════════════════════════════════════════════════════
# EMAIL → PDF
# ══════════════════════════════════════════════════════════════════════════════

# HTML template used to render email messages as printable pages.
# Placeholders: {from_}, {to}, {date}, {subject}, {body}, {attachments}
_EMAIL_TEMPLATE = """\
<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11pt; color: #000; }}
  table {{ width: 100%; border-bottom: 2px solid #444; margin-bottom: 16px; border-collapse: collapse; }}
  td {{ padding: 3px 6px; vertical-align: top; }}
  .lbl {{ color: #555; font-weight: bold; white-space: nowrap; width: 80px; }}
  pre {{ white-space: pre-wrap; word-break: break-word; }}
  .att {{ margin-top: 16px; padding: 8px; border: 1px solid #ccc;
          background: #f9f9f9; font-size: 9pt; }}
</style></head><body>
<table>
  <tr><td class="lbl">From:</td><td>{from_}</td></tr>
  <tr><td class="lbl">To:</td><td>{to}</td></tr>
  <tr><td class="lbl">Date:</td><td>{date}</td></tr>
  <tr><td class="lbl">Subject:</td><td>{subject}</td></tr>
</table>
{body}
{attachments}
</body></html>
"""


def _msg_to_html( msg: Message , logger: logging.Logger ) -> str :
	"""
	Convert a Python email.Message object into an HTML string suitable for
	PDF rendering. Extracts headers, body (HTML preferred, plaintext fallback),
	and lists any attachments with their file sizes.
	"""
	from_ = _decode_header_value( msg.get( "From" ) )
	to = _decode_header_value( msg.get( "To" ) )
	date = _decode_header_value( msg.get( "Date" ) )
	subject = _decode_header_value( msg.get( "Subject" ) ) or "(no subject)"
	logger.debug( "Rendering email to HTML — subject: '%s', from: '%s'" , subject , from_ )

	html_body = plain_body = None
	attachments: list[ tuple[ str , int ] ] = [ ]

	# Walk through all MIME parts to find the body and any attachments
	for part in msg.walk( ) :
		ct = part.get_content_type( )
		disp = str( part.get( "Content-Disposition" , "" ) )
		fname = part.get_filename( )

		# If this part is an attachment, record its name and size
		if fname or "attachment" in disp :
			fname = _decode_header_value( fname ) or f"attachment.{part.get_content_subtype( )}"
			try :
				data = part.get_payload( decode=True ) or b""
				attachments.append( (fname , len( data )) )
				logger.debug( "Found attachment: '%s' (%d bytes)" , fname , len( data ) )
			except Exception as exc :
				logger.warning( "Could not read attachment '%s': %s" , fname , exc )
			continue

		# Prefer HTML body, fall back to plaintext
		if ct == "text/html" and html_body is None :
			try :
				charset = part.get_content_charset( ) or "utf-8"
				html_body = part.get_payload( decode=True ).decode( charset , errors="replace" )
			except Exception as exc :
				logger.warning( "Could not decode HTML body part: %s" , exc )

		elif ct == "text/plain" and plain_body is None :
			try :
				charset = part.get_content_charset( ) or "utf-8"
				plain_body = part.get_payload( decode=True ).decode( charset , errors="replace" )
			except Exception as exc :
				logger.warning( "Could not decode plaintext body part: %s" , exc )

	# Build the final body HTML
	if html_body :
		body = html_body
	elif plain_body :
		# Escape HTML special characters and wrap in <pre> for formatting
		escaped = plain_body.replace( "&" , "&amp;" ).replace( "<" , "&lt;" ).replace( ">" , "&gt;" )
		body = f"<pre>{escaped}</pre>"
	else :
		logger.warning( "Email has no usable body content — subject: '%s'" , subject )
		body = "<p><em>(no body)</em></p>"

	# Build the attachments section if there are any
	att_html = ""
	if attachments :
		items = "<br>".join( f"• {name} ({size:,} bytes)" for name , size in attachments )
		att_html = f'<div class="att"><strong>Attachments ({len( attachments )}):</strong><br>{items}</div>'

	return _EMAIL_TEMPLATE.format(
			from_=from_ , to=to , date=date , subject=subject ,
			body=body , attachments=att_html ,
	)


def _render_msg_to_pdf( msg: Message , dst: Path , logger: logging.Logger ) -> None :
	"""
	Render a single email.Message to a PDF file using WeasyPrint.
	First converts the message to HTML, then prints it to PDF.
	"""
	from weasyprint import HTML , CSS

	html = _msg_to_html( msg , logger )
	HTML( string=html ).write_pdf(
			str( dst ) ,
			stylesheets=[ CSS( string="@page { margin: 20mm 15mm; }" ) ] ,
	)
	logger.debug( "Single-message PDF written: '%s' (%.1f KB)" , dst.name , dst.stat( ).st_size / 1024 )


def _merge_pdfs( pdfs: list[ Path ] , dst: Path , logger: logging.Logger ) -> None :
	"""
	Merge multiple PDF files into a single combined PDF using PyMuPDF (fitz).
	Used when converting multi-message containers (mbox, PST) to a single PDF.
	"""
	import fitz

	logger.debug( "Merging %d individual PDFs into '%s'" , len( pdfs ) , dst.name )
	merged = fitz.open( )
	for p in pdfs :
		with fitz.open( str( p ) ) as doc :
			merged.insert_pdf( doc )
	merged.save( str( dst ) )
	merged.close( )
	logger.debug( "Merged PDF written: '%s' (%.1f KB)" , dst.name , dst.stat( ).st_size / 1024 )


def convert_email_to_pdf( src: Path , logger: logging.Logger ) -> bool :
	"""
	Convert any supported email file to PDF in-place.

	Supported formats:
		.eml / .emlx  — Single email messages
		.mbox / .mbx  — Mailbox containers (multiple messages merged into one PDF)
		.msg          — Microsoft Outlook single message
		.pst / .ost   — Microsoft Outlook data files (all messages merged)

	The original file is deleted on success.

	Args:
		src:    Path to the email file.
		logger: Logger instance.

	Returns:
		True if the conversion succeeded, False otherwise.
	"""
	src = src.resolve( )
	dst = src.with_suffix( ".pdf" )
	ext = src.suffix.lower( )
	logger.info( "Starting email→PDF conversion: '%s' (format: %s)" , src.name , ext )

	try :
		# ── .eml / .emlx — single message files ─────────────────────────
		if ext in (".eml" , ".emlx") :
			raw = src.read_bytes( )

			if ext == ".emlx" :
				# .emlx files have a byte-count on the first line — strip it
				lines = raw.split( b"\n" , 1 )
				if lines[ 0 ].strip( ).isdigit( ) :
					raw = lines[ 1 ]
					logger.debug( "Stripped .emlx byte-count header" )

			msg = email.message_from_bytes( raw , policy=email.policy.compat32 )
			_render_msg_to_pdf( msg , dst , logger )

		# ── .mbox / .mbx — mailbox containers ───────────────────────────
		elif ext in (".mbox" , ".mbx") :
			mbox = mailbox.mbox( str( src ) , create=False )
			msgs = list( mbox.values( ) )

			if not msgs :
				logger.error( "Mailbox file contains no messages: '%s'" , src.name )
				return False

			logger.info( "Mailbox contains %d message(s) — converting each to PDF" , len( msgs ) )

			pdfs , failed = [ ] , 0
			for i , msg in enumerate( msgs ) :
				# Write each message as a separate temporary PDF, then merge
				p = src.parent / f"_mbox_tmp_msg_{i:05d}.pdf"
				try :
					_render_msg_to_pdf( msg , p , logger )
					pdfs.append( p )
				except Exception as exc :
					failed += 1
					logger.warning( "Message %d could not be rendered: %s" , i , exc )

			logger.info(
					"Mailbox conversion results: %d succeeded, %d failed out of %d total" ,
					len( pdfs ) , failed , len( msgs ) ,
			)

			if not pdfs :
				logger.error( "No messages could be converted from '%s'" , src.name )
				return False

			# Merge all individual message PDFs into one file
			_merge_pdfs( pdfs , dst , logger )

			# Clean up temporary per-message PDFs
			for p in pdfs :
				p.unlink( missing_ok=True )

		# ── .msg — Microsoft Outlook single message ──────────────────────
		elif ext == ".msg" :
			import extract_msg

			with extract_msg.openMsg( str( src ) ) as m :
				logger.debug( "Parsed .msg file — subject: '%s', sender: '%s'" , m.subject , m.sender )

				# Build a standard email.Message from the .msg data
				synthetic = email.message_from_string( "" , policy=email.policy.compat32 )
				synthetic[ "From" ] = str( m.sender or "" )
				synthetic[ "To" ] = str( m.to or "" )
				synthetic[ "Date" ] = str( m.date or "" )
				synthetic[ "Subject" ] = str( m.subject or "(no subject)" )

				body = m.htmlBody or m.body
				if body :
					if isinstance( body , bytes ) :
						body = body.decode( "utf-8" , errors="replace" )
					part = email.message.MIMEPart( )
					part.set_type( "text/html" if m.htmlBody else "text/plain" )
					part.set_payload( body )
					synthetic.attach( part )

			_render_msg_to_pdf( synthetic , dst , logger )

		# ── .pst / .ost — Outlook data files ─────────────────────────────
		elif ext in (".pst" , ".ost") :
			import pypff

			def _iter_pst_messages( folder , depth=0 ) -> Iterator :
				"""Recursively yield all messages from a PST folder tree."""
				for i in range( folder.number_of_sub_messages ) :
					try :
						yield folder.get_sub_message( i )
					except Exception as exc :
						logger.warning( "Could not read PST message %d at depth %d: %s" , i , depth , exc )
				for i in range( folder.number_of_sub_folders ) :
					try :
						yield from _iter_pst_messages( folder.get_sub_folder( i ) , depth + 1 )
					except Exception as exc :
						logger.warning( "Could not open PST subfolder %d at depth %d: %s" , i , depth , exc )

			pst = pypff.file( )
			pst.open( str( src ) )

			try :
				root = pst.get_root_folder( )
			except Exception as exc :
				pst.close( )
				logger.error( "Cannot open PST root folder in '%s': %s" , src.name , exc )
				return False

			pdfs , failed = [ ] , 0
			for idx , pst_msg in enumerate( _iter_pst_messages( root ) ) :
				subject = getattr( pst_msg , "subject" , "" ) or "(no subject)"
				p = src.parent / f"_pst_tmp_msg_{idx:06d}.pdf"

				try :
					# Extract body content — prefer HTML, fall back to plaintext
					html_b = pst_msg.html_body
					plain_b = pst_msg.plain_text_body

					if html_b :
						body_html = html_b if isinstance( html_b , str ) else html_b.decode( "utf-8" , errors="replace" )
					elif plain_b :
						t = plain_b if isinstance( plain_b , str ) else plain_b.decode( "utf-8" , errors="replace" )
						body_html = f"<pre>{t.replace( '&' , '&amp;' ).replace( '<' , '&lt;' ).replace( '>' , '&gt;' )}</pre>"
					else :
						body_html = "<p><em>(no body)</em></p>"

					html = _EMAIL_TEMPLATE.format(
							from_=getattr( pst_msg , "sender_name" , "" ) or "" ,
							to=getattr( pst_msg , "display_to" , "" ) or "" ,
							date=str( getattr( pst_msg , "delivery_time" , "" ) or "" ) ,
							subject=subject , body=body_html , attachments="" ,
					)

					from weasyprint import HTML , CSS
					HTML( string=html ).write_pdf(
							str( p ) ,
							stylesheets=[ CSS( string="@page { margin: 20mm 15mm; }" ) ] ,
					)
					pdfs.append( p )

				except Exception as exc :
					failed += 1
					logger.warning( "PST message %d ('%s') could not be rendered: %s" , idx , subject , exc )

			pst.close( )
			logger.info(
					"PST conversion results: %d succeeded, %d failed — '%s'" ,
					len( pdfs ) , failed , src.name ,
			)

			if not pdfs :
				logger.error( "No messages could be extracted from '%s'" , src.name )
				# Clean up any stray temp files
				for p in pdfs :
					p.unlink( missing_ok=True )
				return False

			_merge_pdfs( pdfs , dst , logger )

			# Clean up temporary per-message PDFs
			for p in pdfs :
				p.unlink( missing_ok=True )

		else :
			logger.error(
					"Unsupported email format '%s' for file '%s'. "
					"Supported formats: .eml, .emlx, .mbox, .mbx, .msg, .pst, .ost" ,
					ext , src.name ,
			)
			return False

		# Final check — make sure the output file actually exists
		if not dst.exists( ) :
			logger.error( "Conversion produced no output for '%s'" , src.name )
			return False

		size_kb = dst.stat( ).st_size / 1024
		logger.info( "Email PDF written successfully: '%s' (%.1f KB)" , dst.name , size_kb )
		_delete_original( src , dst , logger )
		return True

	except ImportError as exc :
		logger.error( "Missing dependency for email→PDF conversion: %s" , exc )
		return False

	except Exception as exc :
		logger.error( "Email→PDF conversion failed for '%s': %s" , src.name , exc )
		if dst.exists( ) and dst != src :
			dst.unlink( )
		return False
